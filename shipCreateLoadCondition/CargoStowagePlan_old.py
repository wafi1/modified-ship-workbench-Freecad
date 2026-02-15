#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#***************************************************************************
#*  CargoStowagePlan.py - FLEXIBLE VERSION                                 *
#*  Configurable columns, units, row ranges                                *
#***************************************************************************

import FreeCAD as App
from FreeCAD import Vector
import os
import re

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not installed!")
    print("Install with: pip install openpyxl")


CARGO_COLORS = {
    'container': (0.2, 0.4, 0.8),
    'general':   (0.6, 0.6, 0.6),
    'heavy':     (0.8, 0.2, 0.2),
    'bulk':      (0.8, 0.6, 0.2),
    'vehicle':   (0.3, 0.8, 0.3),
    'default':   (0.7, 0.7, 0.5),
}


class CargoImportConfig:
    """Configuration for Excel import – flexible column and unit mapping."""

    def __init__(self):
        self.col_name        = 'A'
        self.col_description = 'B'
        self.col_mass        = 'C'
        self.col_length      = 'D'
        self.col_width       = 'E'
        self.col_height      = 'F'
        self.col_type        = 'G'

        # row_start : first DATA row  (row 1 is usually the header)
        # row_end   : last row to read; None = read until last filled row
        self.row_start = 2
        self.row_end   = None

        self.mass_unit      = 'kg'
        self.dimension_unit = 'm'

        self.default_description = ""
        self.default_type        = 'default'
        self.default_length      = 1.0
        self.default_width       = 1.0
        self.default_height      = 1.0

    def __str__(self):
        return (
            f"CargoImportConfig:\n"
            f"  Columns : Name={self.col_name}, Mass={self.col_mass}, "
            f"L/W/H={self.col_length}/{self.col_width}/{self.col_height}\n"
            f"  Rows    : {self.row_start} → "
            f"{self.row_end if self.row_end else 'last filled row'}\n"
            f"  Units   : Mass={self.mass_unit}, Dim={self.dimension_unit}"
        )


# ============================================================================
# HELPERS
# ============================================================================

def column_to_index(col):
    """Convert column letter(s) to 1-based index (A=1, B=2, AA=27 …)"""
    if isinstance(col, int):
        return col
    col = col.strip().upper()
    result = 0
    for i, char in enumerate(reversed(col)):
        result += (ord(char) - ord('A') + 1) * (26 ** i)
    return result


def safe_float(value, field_name, row_num, default=None):
    """
    Convert *value* to float safely.

    If the cell contains a label / header text (e.g. 'Mass [kg]') a
    descriptive warning is printed and *default* is returned instead of
    raising an exception.  Unit suffixes embedded in the string
    (e.g. '6.0 m' or '2 500 kg') are stripped automatically.
    """
    if value is None:
        return default

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()

    # Strip trailing unit tokens
    text_clean = re.sub(
        r'\s*(kg|t|lbs?|lb|mm|cm|m|ft|feet|in|inch)\s*$',
        '', text, flags=re.IGNORECASE
    ).strip()

    # Remove thousands separators (space or apostrophe)
    text_clean = text_clean.replace(' ', '').replace("'", '')

    try:
        return float(text_clean)
    except ValueError:
        print(f"  ⚠  Row {row_num} | {field_name}: "
              f"expected a number but got '{value}'"
              + (f" — using default {default}" if default is not None
                 else " — row will be skipped"))
        return default


def convert_mass_to_kg(value, unit):
    conversions = {'kg': 1.0, 't': 1000.0, 'ton': 1000.0,
                   'lbs': 0.453592, 'lb': 0.453592}
    return value * conversions.get(unit.lower(), 1.0)


def convert_dimension_to_m(value, unit):
    conversions = {'m': 1.0, 'meter': 1.0, 'cm': 0.01, 'mm': 0.001,
                   'ft': 0.3048, 'feet': 0.3048, 'in': 0.0254, 'inch': 0.0254}
    return value * conversions.get(unit.lower(), 1.0)


def find_ship_object():
    doc = App.ActiveDocument
    if not doc:
        return None
    for obj in doc.Objects:
        if obj.Label == "Ship" or obj.Name == "Ship":
            return obj
    for obj in doc.Objects:
        if "ship" in obj.Label.lower():
            return obj
    for obj in doc.Objects:
        if hasattr(obj, 'Shape') and obj.Shape:
            bb = obj.Shape.BoundBox
            if bb.XLength > 1000 and bb.YLength > 100:
                return obj
    return None


# ============================================================================
# CORE FUNCTIONS
# ============================================================================

def create_weight_box(doc, name, description, mass_kg,
                      length_m, width_m, height_m,
                      cargo_type='default', position=None):
    """Create a 3-D box as a moveable weight object."""

    box        = doc.addObject("Part::Box", "Cargo")
    box.Length = length_m * 1000   # m → mm (FreeCAD internal unit)
    box.Width  = width_m  * 1000
    box.Height = height_m * 1000
    box.Label  = name

    box.Placement.Base = position if position else Vector(0, 0, 5000)

    box.addProperty("App::PropertyFloat",  "Mass",        "Weight", "Mass in kg").Mass = mass_kg
    box.addProperty("App::PropertyString", "Description", "Weight", "Cargo description").Description = description
    box.addProperty("App::PropertyString", "CargoType",   "Weight", "Type of cargo").CargoType = cargo_type

    # FIX: use .Value to get plain floats from FreeCAD Quantity objects
    # (avoids "Unit mismatch" crash when adding to plain coordinate floats)
    box.addProperty("App::PropertyVector", "COG", "Weight", "Center of Gravity")
    box.COG = Vector(
        box.Placement.Base.x + box.Length.Value / 2,
        box.Placement.Base.y + box.Width.Value  / 2,
        box.Placement.Base.z + box.Height.Value / 2,
    )

    color = CARGO_COLORS.get(cargo_type.lower(), CARGO_COLORS['default'])
    if hasattr(box, 'ViewObject'):
        box.ViewObject.ShapeColor   = color
        box.ViewObject.Transparency = 30

    print(f"  ✓  {name:30s}  {mass_kg:8.0f} kg  "
          f"{length_m:.2f} × {width_m:.2f} × {height_m:.2f} m")
    return box


def update_weight_cog(weight_obj):
    """Recalculate COG from the object's current bounding box."""
    if not hasattr(weight_obj, 'COG'):
        weight_obj.addProperty("App::PropertyVector", "COG", "Weight",
                               "Center of Gravity")
    if hasattr(weight_obj, 'Shape') and weight_obj.Shape:
        bb = weight_obj.Shape.BoundBox
        weight_obj.COG = Vector(
            (bb.XMin + bb.XMax) / 2,
            (bb.YMin + bb.YMax) / 2,
            (bb.ZMin + bb.ZMax) / 2,
        )
        return True
    return False


def add_weight_to_ship(ship, weight_obj):
    if not hasattr(ship, 'Weights'):
        ship.addProperty("App::PropertyStringList", "Weights", "Ship",
                         "List of weight objects")
    weights_list = list(ship.Weights)
    if weight_obj.Name not in weights_list:
        weights_list.append(weight_obj.Name)
        ship.Weights = weights_list
        print("    → Added to Ship.Weights")
        return True
    print("    → Already in Ship.Weights")
    return False


# ============================================================================
# EXCEL READER
# ============================================================================

def read_cargo_excel_flexible(filepath, config, sheet_name=None):
    """
    Read cargo data from Excel with flexible column configuration.

    Rows where numeric fields contain text (headers, sub-totals, …) are
    reported with a clear warning and skipped gracefully.
    """
    if not HAS_OPENPYXL:
        print("ERROR: openpyxl not available!")
        return []
    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}")
        return []

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        print(f"\n{'='*70}")
        print(f"Excel Import : {os.path.basename(filepath)}")
        print(f"Sheet        : {ws.title}")
        print(f"{'='*70}")
        print(config)
        print(f"{'='*70}")

        col_name_idx   = column_to_index(config.col_name)
        col_desc_idx   = column_to_index(config.col_description) if config.col_description else None
        col_mass_idx   = column_to_index(config.col_mass)
        col_length_idx = column_to_index(config.col_length)
        col_width_idx  = column_to_index(config.col_width)
        col_height_idx = column_to_index(config.col_height)
        col_type_idx   = column_to_index(config.col_type) if config.col_type else None

        start_row = config.row_start
        end_row   = config.row_end if config.row_end else ws.max_row

        print(f"\nProcessing rows {start_row} → {end_row} "
              f"({'until last filled row' if not config.row_end else 'fixed end'})\n")

        cargo_list = []
        skipped    = 0

        for row_num in range(start_row, end_row + 1):

            name_raw = ws.cell(row=row_num, column=col_name_idx).value
            if name_raw is None or str(name_raw).strip() == "":
                continue  # empty row → skip silently

            name = str(name_raw).strip()

            # Mass — required; if it's text this is almost certainly a header row
            mass_raw    = ws.cell(row=row_num, column=col_mass_idx).value
            mass_raw_f  = safe_float(mass_raw, "Mass", row_num, default=None)

            if mass_raw_f is None:
                print(f"  ⚠  Row {row_num}: '{name}' skipped "
                      f"(non-numeric mass '{mass_raw}')")
                skipped += 1
                continue

            mass_kg = convert_mass_to_kg(mass_raw_f, config.mass_unit)
            if mass_kg <= 0:
                print(f"  ⚠  Row {row_num}: '{name}' skipped (mass ≤ 0)")
                skipped += 1
                continue

            # Description (optional)
            description = config.default_description
            if col_desc_idx:
                dv = ws.cell(row=row_num, column=col_desc_idx).value
                description = str(dv).strip() if dv else config.default_description

            # Dimensions — use safe_float; fall back to defaults on text
            length_m = convert_dimension_to_m(
                safe_float(ws.cell(row=row_num, column=col_length_idx).value,
                           "Length", row_num, config.default_length),
                config.dimension_unit)
            width_m  = convert_dimension_to_m(
                safe_float(ws.cell(row=row_num, column=col_width_idx).value,
                           "Width",  row_num, config.default_width),
                config.dimension_unit)
            height_m = convert_dimension_to_m(
                safe_float(ws.cell(row=row_num, column=col_height_idx).value,
                           "Height", row_num, config.default_height),
                config.dimension_unit)

            # Cargo type (optional)
            cargo_type = config.default_type
            if col_type_idx:
                tv = ws.cell(row=row_num, column=col_type_idx).value
                cargo_type = str(tv).strip() if tv else config.default_type

            cargo_list.append({
                'name': name, 'description': description,
                'mass_kg': mass_kg, 'length_m': length_m,
                'width_m': width_m, 'height_m': height_m,
                'cargo_type': cargo_type, 'row_num': row_num,
            })

            print(f"  Row {row_num:4d}: {name:30s}  {mass_kg:8.0f} kg  "
                  f"{length_m:.2f}×{width_m:.2f}×{height_m:.2f} m")

        print(f"\n{'='*70}")
        print(f"  ✓  {len(cargo_list)} cargo entries loaded")
        if skipped:
            print(f"  ⚠  {skipped} row(s) skipped")
            print(f"     Tip: check 'From row' — row 1 is usually the header!")
        print(f"{'='*70}\n")

        wb.close()
        return cargo_list

    except Exception as e:
        print(f"ERROR reading Excel file: {e}")
        import traceback
        traceback.print_exc()
        return []


# ============================================================================
# MAIN IMPORT FUNCTION
# ============================================================================

def import_cargo_from_excel(filepath, config=None, sheet_name=None,
                            start_position=None, add_to_ship=True):
    doc = App.ActiveDocument
    if not doc:
        print("ERROR: No active document!")
        return []

    if config is None:
        config = CargoImportConfig()
        print("Using default configuration")

    ship = None
    if add_to_ship:
        ship = find_ship_object()
        if not ship:
            print("WARNING: No Ship object found — cargo created without ship assignment.")

    cargo_list = read_cargo_excel_flexible(filepath, config, sheet_name)
    if not cargo_list:
        print("No cargo data found!")
        return []

    if not start_position:
        if ship and hasattr(ship, 'Shape'):
            bb             = ship.Shape.BoundBox
            start_position = Vector(bb.Center.x, bb.Center.y, bb.ZMax + 500)
        else:
            start_position = Vector(0, 0, 5000)

    print(f"\nCreating 3D objects…\n")
    created_weights = []
    current_pos     = Vector(start_position)
    row_offset      = 0

    for i, cargo in enumerate(cargo_list):
        weight = create_weight_box(
            doc,
            name=cargo['name'], description=cargo['description'],
            mass_kg=cargo['mass_kg'], length_m=cargo['length_m'],
            width_m=cargo['width_m'], height_m=cargo['height_m'],
            cargo_type=cargo['cargo_type'], position=Vector(current_pos),
        )
        if ship and add_to_ship:
            add_weight_to_ship(ship, weight)
        created_weights.append(weight)

        current_pos.x += cargo['length_m'] * 1000 + 500
        if (i + 1) % 5 == 0:
            row_offset   += 1
            current_pos.x = start_position.x
            current_pos.y = start_position.y + row_offset * 3000

    doc.recompute()

    print(f"\n{'='*70}")
    print(f"  ✓  {len(created_weights)} cargo object(s) created")
    if ship and add_to_ship:
        print(f"  ✓  All added to ship '{ship.Label}'")
    print(f"\n  → Objects can be repositioned in the 3D view")
    print(f"  → Click 'Update COG' afterwards")
    print(f"{'='*70}\n")
    return created_weights


def update_all_cargo_cogs():
    doc = App.ActiveDocument
    if not doc:
        return
    updated = 0
    for obj in doc.Objects:
        if hasattr(obj, 'Mass') and hasattr(obj, 'CargoType'):
            if update_weight_cog(obj):
                updated += 1
    doc.recompute()
    print(f"\n  ✓  COG updated for {updated} cargo object(s)\n")


def create_cargo_group(doc, cargo_objects, group_name="Cargo"):
    group = doc.addObject("App::DocumentObjectGroup", group_name)
    group.Label = group_name
    for obj in cargo_objects:
        group.addObject(obj)
    print(f"  ✓  Group '{group_name}' created ({len(cargo_objects)} objects)")
    return group


# ============================================================================
# PRE-DEFINED CONFIGURATIONS
# ============================================================================

def get_standard_config():
    c = CargoImportConfig()
    c.col_name='A'; c.col_description='B'; c.col_mass='C'
    c.col_length='D'; c.col_width='E'; c.col_height='F'; c.col_type='G'
    c.row_start=2; c.row_end=None; c.mass_unit='kg'; c.dimension_unit='m'
    return c


def get_custom_config(name_col='A', mass_col='C', length_col='D',
                      width_col='E', height_col='F',
                      start_row=2, end_row=None,
                      mass_unit='kg', dim_unit='m'):
    c = CargoImportConfig()
    c.col_name=name_col; c.col_mass=mass_col; c.col_length=length_col
    c.col_width=width_col; c.col_height=height_col
    c.row_start=start_row; c.row_end=end_row
    c.mass_unit=mass_unit; c.dimension_unit=dim_unit
    return c
