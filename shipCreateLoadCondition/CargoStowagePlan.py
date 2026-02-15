#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#***************************************************************************
#*  CargoStowagePlan.py - WITH INTEGRATED 3D TEXT LABELS                   *
#*  Single-object cargo boxes with engraved or raised 3D text              *
#*  FIXED: Unit mismatch in COG expressions resolved                       *
#***************************************************************************

import FreeCAD as App
from FreeCAD import Vector
import os
import re

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not installed!")

try:
    import Draft
    HAS_DRAFT = True
except ImportError:
    HAS_DRAFT = False

try:
    import Part
    HAS_PART = True
except ImportError:
    HAS_PART = False


CARGO_COLORS = {
    'container': (0.2, 0.4, 0.8),
    'general':   (0.6, 0.6, 0.6),
    'heavy':     (0.8, 0.2, 0.2),
    'bulk':      (0.8, 0.6, 0.2),
    'vehicle':   (0.3, 0.8, 0.3),
    'default':   (0.7, 0.7, 0.5),
}


class CargoImportConfig:
    """Configuration for Excel import."""

    def __init__(self):
        self.col_name        = 'A'
        self.col_description = 'B'
        self.col_mass        = 'C'
        self.col_length      = 'D'
        self.col_width       = 'E'
        self.col_height      = 'F'
        self.col_type        = 'G'
        self.row_start = 2
        self.row_end   = None
        self.mass_unit      = 'kg'
        self.dimension_unit = 'm'
        self.default_description = ""
        self.default_type        = 'default'
        self.default_length      = 1.0
        self.default_width       = 1.0
        self.default_height      = 1.0
        self.use_3d_text       = True
        self.engrave_text      = False
        self.font_file         = None

    def __str__(self):
        return (
            f"CargoImportConfig:\n"
            f"  Columns : Name={self.col_name}, Mass={self.col_mass}\n"
            f"  Rows    : {self.row_start} → "
            f"{self.row_end if self.row_end else 'last'}\n"
            f"  3D Text : {'Enabled (' + ('engraved' if self.engrave_text else 'raised') + ')' if self.use_3d_text else 'Disabled'}"
        )


def column_to_index(col):
    """Convert column letter to 1-based index."""
    if isinstance(col, int):
        return col
    col = col.strip().upper()
    result = 0
    for i, char in enumerate(reversed(col)):
        result += (ord(char) - ord('A') + 1) * (26 ** i)
    return result


def safe_float(value, field_name, row_num, default=None):
    """Convert value to float safely."""
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    text_clean = re.sub(r'\s*(kg|t|lbs?|lb|mm|cm|m|ft|feet|in|inch)\s*$',
                        '', text, flags=re.IGNORECASE).strip()
    text_clean = text_clean.replace(' ', '').replace("'", '')
    try:
        return float(text_clean)
    except ValueError:
        print(f"  ⚠  Row {row_num} | {field_name}: expected number, got '{value}'")
        return default


def convert_mass_to_kg(value, unit):
    conversions = {'kg': 1.0, 't': 1000.0, 'ton': 1000.0,
                   'lbs': 0.453592, 'lb': 0.453592}
    return value * conversions.get(unit.lower(), 1.0)


def convert_dimension_to_m(value, unit):
    conversions = {'m': 1.0, 'cm': 0.01, 'mm': 0.001,
                   'ft': 0.3048, 'in': 0.0254}
    return value * conversions.get(unit.lower(), 1.0)


def find_ship_object():
    doc = App.ActiveDocument
    if not doc:
        return None
    for obj in doc.Objects:
        if obj.Label == "Ship" or obj.Name == "Ship":
            return obj
    for obj in doc.Objects:
        if "ship" in obj.Label.lower():
            return obj
    return None


def find_font_file():
    """Try to find a suitable font file."""
    import platform
    system = platform.system()
    
    if system == "Linux":
        paths = [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
        ]
    elif system == "Windows":
        paths = [
            "C:/Windows/Fonts/arialbd.ttf",
            "C:/Windows/Fonts/arial.ttf",
            "C:/Windows/Fonts/calibrib.ttf",
        ]
    elif system == "Darwin":
        paths = [
            "/System/Library/Fonts/Helvetica.ttc",
            "/Library/Fonts/Arial Bold.ttf",
        ]
    else:
        paths = []
    
    for path in paths:
        if os.path.exists(path):
            return path
    return None


def create_weight_box_with_3d_label(doc, name, description, mass_kg,
                                     length_m, width_m, height_m,
                                     cargo_type='default', position=None,
                                     engrave=False, font_file=None):
    """Create cargo box with integrated 3D text - FIXED UNIT MISMATCH."""
    
    if not HAS_PART:
        return create_weight_box_simple(doc, name, description, mass_kg,
                                        length_m, width_m, height_m,
                                        cargo_type, position)
    
    L = length_m * 1000
    W = width_m * 1000
    H = height_m * 1000
    
    box_shape = Part.makeBox(L, W, H)
    
    min_dim = min(L, W, H)
    font_size = max(min_dim * 0.12, 80)
    font_size = min(font_size, 400)
    
    mass_t = mass_kg / 1000.0
    line1 = name[:20] if len(name) > 20 else name
    line2 = f"{mass_t:.1f}t"
    line3 = f"{length_m:.1f}x{width_m:.1f}x{height_m:.1f}"
    
    if font_file is None:
        font_file = find_font_file()
    
    if not font_file or not os.path.exists(font_file):
        print(f"  ⚠  Font not found, box without text: {name}")
        final_shape = box_shape
    else:
        try:
            temp_objects = []
            text_shapes = []
            y_offset = 0
            line_spacing = font_size * 1.3
            
            for line_text in [line1, line2, line3]:
                if not line_text.strip():
                    continue
                
                ss = Draft.makeShapeString(
                    String=line_text,
                    FontFile=font_file,
                    Size=font_size,
                    Tracking=0
                )
                
                if ss and ss.Shape and not ss.Shape.isNull():
                    text_shape_copy = ss.Shape.copy()
                    text_bb = text_shape_copy.BoundBox
                    text_width = text_bb.XLength
                    
                    x_center = (L - text_width) / 2
                    y_pos = (W / 2) + y_offset
                    z_pos = H if engrave else H + 2
                    
                    placement = App.Placement(
                        App.Vector(x_center, y_pos, z_pos),
                        App.Rotation()
                    )
                    
                    positioned_shape = text_shape_copy.copy()
                    positioned_shape.Placement = placement
                    
                    extrude_depth = font_size * 0.25
                    
                    if engrave:
                        text_solid = positioned_shape.extrude(App.Vector(0, 0, -extrude_depth))
                    else:
                        text_solid = positioned_shape.extrude(App.Vector(0, 0, extrude_depth))
                    
                    text_shapes.append(text_solid)
                    temp_objects.append(ss.Name)
                    y_offset -= line_spacing
            
            # Delete temporary objects immediately
            for obj_name in temp_objects:
                try:
                    doc.removeObject(obj_name)
                except:
                    pass
            
            if text_shapes:
                combined_text = text_shapes[0]
                for ts in text_shapes[1:]:
                    combined_text = combined_text.fuse(ts)
                
                if engrave:
                    final_shape = box_shape.cut(combined_text)
                else:
                    final_shape = box_shape.fuse(combined_text)
            else:
                final_shape = box_shape
            
        except Exception as e:
            print(f"  ⚠  3D text failed for {name}: {e}")
            final_shape = box_shape
    
    cargo = doc.addObject("Part::Feature", "Cargo")
    cargo.Shape = final_shape
    cargo.Label = name
    
    if position is None:
        position = Vector(0, 0, 5000)
    cargo.Placement.Base = position
    
    cargo.addProperty("App::PropertyFloat", "Mass", "Weight", "Mass in kg").Mass = mass_kg
    
    # FIXED: Use PropertyLength instead of PropertyFloat for dimensions
    cargo.addProperty("App::PropertyLength", "Length", "Dimensions", "Length in mm").Length = L
    cargo.addProperty("App::PropertyLength", "Width", "Dimensions", "Width in mm").Width = W
    cargo.addProperty("App::PropertyLength", "Height", "Dimensions", "Height in mm").Height = H
    
    cargo.addProperty("App::PropertyString", "Description", "Weight", "Description").Description = description
    cargo.addProperty("App::PropertyString", "CargoType", "Weight", "Type").CargoType = cargo_type
    
    # FIXED: COG with expressions - no unit mismatch because both are Length
    cargo.addProperty("App::PropertyVector", "COG", "Weight", "Center of Gravity")
    try:
        cargo.setExpression('COG.x', 'Placement.Base.x + Length / 2')
        cargo.setExpression('COG.y', 'Placement.Base.y + Width / 2')
        cargo.setExpression('COG.z', 'Placement.Base.z + Height / 2')
    except Exception as e:
        print(f"  ⚠  Could not bind COG: {e}")
        cargo.COG = Vector(position.x + L/2, position.y + W/2, position.z + H/2)
    
    color = CARGO_COLORS.get(cargo_type.lower(), CARGO_COLORS['default'])
    if hasattr(cargo, 'ViewObject'):
        cargo.ViewObject.ShapeColor = color
        cargo.ViewObject.Transparency = 20
    
    print(f"  ✓  {name:30s}  {mass_kg:8.0f} kg  "
          f"{length_m:.2f}×{width_m:.2f}×{height_m:.2f}m  "
          f"({'engr' if engrave else 'top'})")
    
    return cargo


def create_weight_box_simple(doc, name, description, mass_kg,
                             length_m, width_m, height_m,
                             cargo_type='default', position=None):
    """Simple box without 3D text - FIXED UNIT MISMATCH."""
    
    box = doc.addObject("Part::Box", "Cargo")
    box.Length = length_m * 1000
    box.Width  = width_m  * 1000
    box.Height = height_m * 1000
    box.Label  = name
    
    box.Placement.Base = position if position else Vector(0, 0, 5000)
    
    # Add custom properties (Length/Width/Height already exist in Part::Box)
    box.addProperty("App::PropertyFloat", "Mass", "Weight", "Mass in kg").Mass = mass_kg
    box.addProperty("App::PropertyString", "Description", "Weight", "Description").Description = description
    box.addProperty("App::PropertyString", "CargoType", "Weight", "Type").CargoType = cargo_type
    
    # FIXED: COG with expressions - Part::Box Length/Width/Height are PropertyLength
    box.addProperty("App::PropertyVector", "COG", "Weight", "Center of Gravity")
    try:
        box.setExpression('COG.x', 'Placement.Base.x + Length / 2')
        box.setExpression('COG.y', 'Placement.Base.y + Width / 2')
        box.setExpression('COG.z', 'Placement.Base.z + Height / 2')
    except Exception as e:
        print(f"  ⚠  Could not bind COG: {e}")
        box.COG = Vector(
            box.Placement.Base.x + box.Length / 2,
            box.Placement.Base.y + box.Width / 2,
            box.Placement.Base.z + box.Height / 2)
    
    color = CARGO_COLORS.get(cargo_type.lower(), CARGO_COLORS['default'])
    if hasattr(box, 'ViewObject'):
        box.ViewObject.ShapeColor = color
        box.ViewObject.Transparency = 30
    
    print(f"  ✓  {name:30s}  {mass_kg:8.0f} kg  {length_m:.2f}×{width_m:.2f}×{height_m:.2f}m")
    return box


def add_weight_to_ship(ship, weight_obj):
    if not hasattr(ship, 'Weights'):
        ship.addProperty("App::PropertyStringList", "Weights", "Ship", "Weights")
    weights_list = list(ship.Weights)
    if weight_obj.Name not in weights_list:
        weights_list.append(weight_obj.Name)
        ship.Weights = weights_list
        print("    → Added to Ship.Weights")
        return True
    return False


def read_cargo_excel_flexible(filepath, config, sheet_name=None):
    """Read cargo from Excel."""
    if not HAS_OPENPYXL:
        print("ERROR: openpyxl not available!")
        return []
    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}")
        return []

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        print(f"\n{'='*70}")
        print(f"Excel Import : {os.path.basename(filepath)}")
        print(f"Sheet        : {ws.title}")
        print(f"{'='*70}")
        print(config)
        print(f"{'='*70}")

        col_name_idx   = column_to_index(config.col_name)
        col_desc_idx   = column_to_index(config.col_description) if config.col_description else None
        col_mass_idx   = column_to_index(config.col_mass)
        col_length_idx = column_to_index(config.col_length)
        col_width_idx  = column_to_index(config.col_width)
        col_height_idx = column_to_index(config.col_height)
        col_type_idx   = column_to_index(config.col_type) if config.col_type else None

        start_row = config.row_start
        end_row   = config.row_end if config.row_end else ws.max_row

        print(f"\nProcessing rows {start_row} → {end_row}\n")

        cargo_list = []
        skipped    = 0

        for row_num in range(start_row, end_row + 1):
            name_raw = ws.cell(row=row_num, column=col_name_idx).value
            if not name_raw or str(name_raw).strip() == "":
                continue

            name = str(name_raw).strip()

            mass_raw = ws.cell(row=row_num, column=col_mass_idx).value
            mass_raw_f = safe_float(mass_raw, "Mass", row_num, default=None)

            if mass_raw_f is None:
                skipped += 1
                continue

            mass_kg = convert_mass_to_kg(mass_raw_f, config.mass_unit)
            if mass_kg <= 0:
                skipped += 1
                continue

            description = config.default_description
            if col_desc_idx:
                dv = ws.cell(row=row_num, column=col_desc_idx).value
                description = str(dv).strip() if dv else config.default_description

            length_m = convert_dimension_to_m(
                safe_float(ws.cell(row=row_num, column=col_length_idx).value,
                           "Length", row_num, config.default_length),
                config.dimension_unit)
            width_m = convert_dimension_to_m(
                safe_float(ws.cell(row=row_num, column=col_width_idx).value,
                           "Width", row_num, config.default_width),
                config.dimension_unit)
            height_m = convert_dimension_to_m(
                safe_float(ws.cell(row=row_num, column=col_height_idx).value,
                           "Height", row_num, config.default_height),
                config.dimension_unit)

            cargo_type = config.default_type
            if col_type_idx:
                tv = ws.cell(row=row_num, column=col_type_idx).value
                cargo_type = str(tv).strip() if tv else config.default_type

            cargo_list.append({
                'name': name, 'description': description,
                'mass_kg': mass_kg, 'length_m': length_m,
                'width_m': width_m, 'height_m': height_m,
                'cargo_type': cargo_type, 'row_num': row_num,
            })

        print(f"\n{'='*70}")
        print(f"  ✓  {len(cargo_list)} cargo entries loaded")
        if skipped:
            print(f"  ⚠  {skipped} row(s) skipped")
        print(f"{'='*70}\n")

        wb.close()
        return cargo_list

    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        return []


def import_cargo_from_excel(filepath, config=None, sheet_name=None,
                            start_position=None, add_to_ship=True):
    """Import cargo from Excel with 3D text."""
    doc = App.ActiveDocument
    if not doc:
        print("ERROR: No active document!")
        return []

    if config is None:
        config = CargoImportConfig()

    ship = None
    if add_to_ship:
        ship = find_ship_object()

    cargo_list = read_cargo_excel_flexible(filepath, config, sheet_name)
    if not cargo_list:
        return []

    if not start_position:
        if ship and hasattr(ship, 'Shape'):
            bb = ship.Shape.BoundBox
            start_position = Vector(bb.Center.x, bb.Center.y, bb.ZMax + 500)
        else:
            start_position = Vector(0, 0, 5000)

    print(f"\nCreating 3D objects...\n")
    created_weights = []
    current_pos = Vector(start_position)
    row_offset = 0

    for i, cargo in enumerate(cargo_list):
        if config.use_3d_text:
            weight = create_weight_box_with_3d_label(
                doc, cargo['name'], cargo['description'], cargo['mass_kg'],
                cargo['length_m'], cargo['width_m'], cargo['height_m'],
                cargo['cargo_type'], Vector(current_pos),
                config.engrave_text, config.font_file)
        else:
            weight = create_weight_box_simple(
                doc, cargo['name'], cargo['description'], cargo['mass_kg'],
                cargo['length_m'], cargo['width_m'], cargo['height_m'],
                cargo['cargo_type'], Vector(current_pos))
        
        if ship and add_to_ship:
            add_weight_to_ship(ship, weight)
        created_weights.append(weight)

        current_pos.x += cargo['length_m'] * 1000 + 500
        if (i + 1) % 5 == 0:
            row_offset += 1
            current_pos.x = start_position.x
            current_pos.y = start_position.y + row_offset * 3000

    # Clean up leftover ShapeString objects
    for obj in doc.Objects:
        if 'ShapeString' in str(obj.TypeId):
            try:
                doc.removeObject(obj.Name)
            except:
                pass

    doc.recompute()

    print(f"\n{'='*70}")
    print(f"  ✓  {len(created_weights)} cargo created")
    if config.use_3d_text:
        print(f"  ✓  Text {'engraved' if config.engrave_text else 'on top'}")
    print(f"  ✓  Single solid objects - move freely")
    print(f"  ✓  COG follows position automatically")
    if ship and add_to_ship:
        print(f"  ✓  Added to ship '{ship.Label}'")
    print(f"{'='*70}\n")
    return created_weights


# ============================================================================
# CONFIGURATIONS
# ============================================================================

def get_standard_config():
    """Standard with raised 3D text."""
    c = CargoImportConfig()
    c.col_name='A'; c.col_description='B'; c.col_mass='C'
    c.col_length='D'; c.col_width='E'; c.col_height='F'; c.col_type='G'
    c.row_start=2; c.row_end=None
    c.mass_unit='kg'; c.dimension_unit='m'
    c.use_3d_text = True
    c.engrave_text = False
    return c


def get_engraved_config():
    """With engraved text."""
    c = get_standard_config()
    c.engrave_text = True
    return c


def get_simple_config():
    """Without 3D text."""
    c = get_standard_config()
    c.use_3d_text = False
    return c


def get_custom_config(name_col='A', mass_col='C', length_col='D',
                      width_col='E', height_col='F',
                      start_row=2, end_row=None,
                      mass_unit='kg', dim_unit='m',
                      use_3d_text=True, engrave_text=False,
                      font_file=None):
    """Custom configuration."""
    c = CargoImportConfig()
    c.col_name=name_col; c.col_mass=mass_col
    c.col_length=length_col; c.col_width=width_col; c.col_height=height_col
    c.row_start=start_row; c.row_end=end_row
    c.mass_unit=mass_unit; c.dimension_unit=dim_unit
    c.use_3d_text = use_3d_text
    c.engrave_text = engrave_text
    c.font_file = font_file
    return c


def update_all_cargo_cogs():
    """Update COG for all cargo (usually not needed with expressions)."""
    doc = App.ActiveDocument
    if not doc:
        return
    count = 0
    for obj in doc.Objects:
        if hasattr(obj, 'Mass') and hasattr(obj, 'COG'):
            if hasattr(obj, 'Shape') and obj.Shape:
                bb = obj.Shape.BoundBox
                obj.COG = Vector(
                    (bb.XMin + bb.XMax) / 2,
                    (bb.YMin + bb.YMax) / 2,
                    (bb.ZMin + bb.ZMax) / 2)
                count += 1
    if count > 0:
        doc.recompute()
        print(f"\n  ✓  COG updated for {count} cargo objects\n")


def create_cargo_group(doc, cargo_objects, group_name="Cargo"):
    """Create group for cargo objects."""
    group = doc.addObject("App::DocumentObjectGroup", group_name)
    group.Label = group_name
    for obj in cargo_objects:
        group.addObject(obj)
    print(f"  ✓  Group '{group_name}' created with {len(cargo_objects)} objects")
    return group


# ============================================================================
# GUI COMMAND CLASS
# ============================================================================

class CargoImportCommand:
    """FreeCAD GUI Command for cargo import."""
    
    def GetResources(self):
        import os
        icon_path = os.path.join(os.path.dirname(__file__), "..", "resources", "icons", "ship_cargo.svg")
        return {
            'Pixmap': icon_path if os.path.exists(icon_path) else '',
            'MenuText': "Import Cargo from Excel",
            'ToolTip': "Import cargo boxes with 3D labels from Excel file"
        }
    
    def Activated(self):
        """Called when command is activated."""
        from PySide import QtGui
        
        # File dialog
        filepath, _ = QtGui.QFileDialog.getOpenFileName(
            None,
            "Select Excel File",
            "",
            "Excel Files (*.xlsx *.xls);;All Files (*.*)"
        )
        
        if not filepath:
            return
        
        # Use standard config
        config = get_standard_config()
        
        # Import
        try:
            import_cargo_from_excel(filepath, config=config)
        except Exception as e:
            print(f"ERROR: {e}")
            import traceback
            traceback.print_exc()
    
    def IsActive(self):
        """Command is active when document exists."""
        return App.ActiveDocument is not None


# Register command
try:
    import FreeCADGui as Gui
    Gui.addCommand('Ship_ImportCargo', CargoImportCommand())
except:
    pass


"""
USAGE:

1. Standard (raised text):
   config = get_standard_config()
   import_cargo_from_excel('/path/file.xlsx', config=config)

2. Engraved text:
   config = get_engraved_config()
   import_cargo_from_excel('/path/file.xlsx', config=config)

3. No text:
   config = get_simple_config()
   import_cargo_from_excel('/path/file.xlsx', config=config)

FEATURES:
✓ Single solid object (box + text fused)
✓ Move freely - no dependencies
✓ COG follows automatically (FIXED unit mismatch)
✓ No ghost boxes
"""
