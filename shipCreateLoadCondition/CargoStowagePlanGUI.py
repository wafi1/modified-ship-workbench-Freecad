#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#***************************************************************************
#*  CargoStowagePlanGUI.py – GUI for Cargo Import                         *
#*  Simple operation, preview, live updates                               *
#***************************************************************************

import FreeCAD as App
import FreeCADGui as Gui
from PySide2 import QtWidgets, QtCore, QtGui
import os

# Relative import — CargoStowagePlan.py lives in the same sub-package
from .CargoStowagePlan import (
    CargoImportConfig, get_standard_config, get_custom_config,
    import_cargo_from_excel, update_all_cargo_cogs,
    HAS_OPENPYXL, find_ship_object
)


class CargoImportDialog(QtWidgets.QDialog):
    """Main dialog for cargo import"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Cargo Stowage Plan Import")
        self.setMinimumSize(820, 640)

        self.config          = get_standard_config()
        self.created_weights = []
        self.preview_data    = []

        self.setup_ui()
        self.load_defaults()

        self.preview_timer = QtCore.QTimer()
        self.preview_timer.setSingleShot(True)
        self.preview_timer.timeout.connect(self.update_preview)

    # ======================================================================
    def setup_ui(self):
        main_layout = QtWidgets.QVBoxLayout()

        header = QtWidgets.QLabel("📦  CARGO STOWAGE PLAN IMPORT")
        header.setStyleSheet("""
            QLabel {
                font-size: 18px; font-weight: bold; padding: 10px;
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #3498db, stop:1 #2ecc71);
                border-radius: 5px; color: white;
            }
        """)
        main_layout.addWidget(header)

        tabs = QtWidgets.QTabWidget()
        tab1 = QtWidgets.QWidget(); self.setup_tab1(tab1); tabs.addTab(tab1, "📄 File")
        tab2 = QtWidgets.QWidget(); self.setup_tab2(tab2); tabs.addTab(tab2, "🗂️ Columns")
        tab3 = QtWidgets.QWidget(); self.setup_tab3(tab3); tabs.addTab(tab3, "⚙️ Settings")
        main_layout.addWidget(tabs)

        main_layout.addWidget(self.create_preview_section())
        main_layout.addWidget(self.create_button_section())

        self.status_bar = QtWidgets.QStatusBar()
        self.status_bar.showMessage("Ready")
        main_layout.addWidget(self.status_bar)

        self.setLayout(main_layout)

    # ======================================================================
    def setup_tab1(self, parent):
        layout = QtWidgets.QVBoxLayout()

        # --- File selection ---
        file_group  = QtWidgets.QGroupBox("Excel File")
        file_layout = QtWidgets.QHBoxLayout()
        self.file_path = QtWidgets.QLineEdit()
        self.file_path.setPlaceholderText("Select an Excel file…")
        file_layout.addWidget(self.file_path)
        browse_btn = QtWidgets.QPushButton("📂 Browse")
        browse_btn.clicked.connect(self.browse_file)
        file_layout.addWidget(browse_btn)
        file_group.setLayout(file_layout)
        layout.addWidget(file_group)

        # --- Sheet selection ---
        sheet_group  = QtWidgets.QGroupBox("Worksheet")
        sheet_layout = QtWidgets.QHBoxLayout()
        self.sheet_combo = QtWidgets.QComboBox()
        self.sheet_combo.addItem("(Automatic)", None)
        sheet_layout.addWidget(self.sheet_combo, 3)
        sheet_layout.addWidget(QtWidgets.QLabel("or name:"))
        self.sheet_name = QtWidgets.QLineEdit()
        self.sheet_name.setPlaceholderText("Sheet1")
        sheet_layout.addWidget(self.sheet_name, 2)
        reload_btn = QtWidgets.QPushButton("🔄 Load sheets")
        reload_btn.clicked.connect(self.load_sheets)
        sheet_layout.addWidget(reload_btn)
        sheet_group.setLayout(sheet_layout)
        layout.addWidget(sheet_group)

        # --- Row range ---
        row_group  = QtWidgets.QGroupBox("Row Range")
        row_layout = QtWidgets.QGridLayout()

        row_layout.addWidget(QtWidgets.QLabel("From row:"), 0, 0)
        self.start_row = QtWidgets.QSpinBox()
        self.start_row.setRange(1, 100000)
        self.start_row.setValue(2)          # row 1 is normally the header
        self.start_row.setToolTip("First data row (row 1 is the header row)")
        row_layout.addWidget(self.start_row, 0, 1)

        row_layout.addWidget(QtWidgets.QLabel("To row:"), 1, 0)
        self.end_row = QtWidgets.QSpinBox()
        # FIX: minimum = 0 so that value 0 means "read until last filled row"
        self.end_row.setRange(0, 100000)
        self.end_row.setValue(0)            # 0 = until last filled row
        self.end_row.setSpecialValueText("Last filled row")
        self.end_row.setToolTip(
            "Last row to read.\n"
            "Set to 0 (= 'Last filled row') to read the whole sheet.")
        row_layout.addWidget(self.end_row, 1, 1)

        hint = QtWidgets.QLabel(
            "ℹ  Row 1 usually contains column headers — start from row 2.\n"
            "   Set 'To row' to 0 to read until the last filled row."
        )
        hint.setStyleSheet("color: #555; font-style: italic; font-size: 11px;")
        row_layout.addWidget(hint, 2, 0, 1, 2)

        row_group.setLayout(row_layout)
        layout.addWidget(row_group)

        self.auto_preview = QtWidgets.QCheckBox("Automatic preview on changes")
        self.auto_preview.setChecked(True)
        layout.addWidget(self.auto_preview)

        layout.addStretch()
        parent.setLayout(layout)

    # ======================================================================
    def setup_tab2(self, parent):
        layout = QtWidgets.QGridLayout()

        layout.addWidget(QtWidgets.QLabel("Field"),   0, 0)
        layout.addWidget(QtWidgets.QLabel("Column"),  0, 1)
        layout.addWidget(QtWidgets.QLabel("Example"), 0, 2)

        self.column_widgets = {}
        columns = [
            ('Name',        'col_name',        'A', 'Cargo001'),
            ('Description', 'col_description', 'B', 'General goods'),
            ('Mass',        'col_mass',        'C', '2500'),
            ('Length',      'col_length',      'D', '6.0'),
            ('Width',       'col_width',       'E', '2.4'),
            ('Height',      'col_height',      'F', '2.6'),
            ('Type',        'col_type',        'G', 'container'),
        ]

        for i, (label, field, default, example) in enumerate(columns, 1):
            layout.addWidget(QtWidgets.QLabel(label), i, 0)
            le = QtWidgets.QLineEdit(default)
            le.setMaximumWidth(50)
            le.textChanged.connect(self.on_column_changed)
            layout.addWidget(le, i, 1)
            layout.addWidget(QtWidgets.QLabel(f"e.g. '{example}'"), i, 2)
            self.column_widgets[field] = le

        preset_layout = QtWidgets.QHBoxLayout()
        std_btn = QtWidgets.QPushButton("⚡ Standard Layout")
        std_btn.clicked.connect(self.load_standard_preset)
        preset_layout.addWidget(std_btn)
        cus_btn = QtWidgets.QPushButton("🎯 Custom Layout")
        cus_btn.clicked.connect(self.load_custom_preset)
        preset_layout.addWidget(cus_btn)
        layout.addLayout(preset_layout, len(columns) + 1, 0, 1, 3)

        parent.setLayout(layout)

    # ======================================================================
    def setup_tab3(self, parent):
        layout = QtWidgets.QVBoxLayout()

        units_group  = QtWidgets.QGroupBox("Units")
        units_layout = QtWidgets.QGridLayout()
        units_layout.addWidget(QtWidgets.QLabel("Mass:"), 0, 0)
        self.mass_unit = QtWidgets.QComboBox()
        self.mass_unit.addItems(['kg', 't (tonnes)', 'lbs (pounds)'])
        units_layout.addWidget(self.mass_unit, 0, 1)
        units_layout.addWidget(QtWidgets.QLabel("Dimensions:"), 1, 0)
        self.dim_unit = QtWidgets.QComboBox()
        self.dim_unit.addItems(['m', 'cm', 'mm', 'ft (feet)', 'in (inches)'])
        units_layout.addWidget(self.dim_unit, 1, 1)
        units_group.setLayout(units_layout)
        layout.addWidget(units_group)

        defaults_group  = QtWidgets.QGroupBox("Default values (used when a cell is empty)")
        defaults_layout = QtWidgets.QGridLayout()
        defaults_layout.addWidget(QtWidgets.QLabel("Description:"), 0, 0)
        self.default_desc = QtWidgets.QLineEdit("")
        defaults_layout.addWidget(self.default_desc, 0, 1)
        defaults_layout.addWidget(QtWidgets.QLabel("Type:"), 1, 0)
        self.default_type = QtWidgets.QComboBox()
        self.default_type.addItems(['default','container','general','heavy','bulk','vehicle'])
        defaults_layout.addWidget(self.default_type, 1, 1)
        defaults_layout.addWidget(QtWidgets.QLabel("Length:"), 2, 0)
        self.default_length = QtWidgets.QDoubleSpinBox()
        self.default_length.setRange(0.1, 100.0); self.default_length.setValue(1.0)
        self.default_length.setSuffix(" m")
        defaults_layout.addWidget(self.default_length, 2, 1)
        defaults_layout.addWidget(QtWidgets.QLabel("Width:"), 3, 0)
        self.default_width = QtWidgets.QDoubleSpinBox()
        self.default_width.setRange(0.1, 100.0); self.default_width.setValue(1.0)
        self.default_width.setSuffix(" m")
        defaults_layout.addWidget(self.default_width, 3, 1)
        defaults_layout.addWidget(QtWidgets.QLabel("Height:"), 4, 0)
        self.default_height = QtWidgets.QDoubleSpinBox()
        self.default_height.setRange(0.1, 100.0); self.default_height.setValue(1.0)
        self.default_height.setSuffix(" m")
        defaults_layout.addWidget(self.default_height, 4, 1)
        defaults_group.setLayout(defaults_layout)
        layout.addWidget(defaults_group)

        ship_group  = QtWidgets.QGroupBox("Ship")
        ship_layout = QtWidgets.QVBoxLayout()
        self.add_to_ship = QtWidgets.QCheckBox("Automatically add cargo to ship")
        self.add_to_ship.setChecked(True)
        ship_layout.addWidget(self.add_to_ship)
        ship_info = QtWidgets.QLabel("(Requires a Ship object in the document)")
        ship_info.setStyleSheet("color: #666; font-style: italic;")
        ship_layout.addWidget(ship_info)
        ship_group.setLayout(ship_layout)
        layout.addWidget(ship_group)

        layout.addStretch()
        parent.setLayout(layout)

    # ======================================================================
    def create_preview_section(self):
        group  = QtWidgets.QGroupBox("📊 Preview")
        layout = QtWidgets.QVBoxLayout()
        self.preview_table = QtWidgets.QTableWidget()
        self.preview_table.setColumnCount(3)
        self.preview_table.setHorizontalHeaderLabels(
            ["Row", "Name", "Mass (raw value)"])
        self.preview_table.horizontalHeader().setStretchLastSection(True)
        self.preview_table.setAlternatingRowColors(True)
        layout.addWidget(self.preview_table)

        ctrl = QtWidgets.QHBoxLayout()
        self.preview_count = QtWidgets.QLabel("0 entries loaded")
        ctrl.addWidget(self.preview_count)
        ctrl.addStretch()
        pb = QtWidgets.QPushButton("🔍 Refresh preview")
        pb.clicked.connect(self.update_preview)
        ctrl.addWidget(pb)
        layout.addLayout(ctrl)
        group.setLayout(layout)
        return group

    # ======================================================================
    def create_button_section(self):
        widget = QtWidgets.QWidget()
        layout = QtWidgets.QHBoxLayout()

        self.import_btn = QtWidgets.QPushButton("🚀  START IMPORT")
        self.import_btn.setStyleSheet("""
            QPushButton {
                background-color:#27ae60; color:white; font-weight:bold;
                padding:12px; font-size:14px; border-radius:5px;
            }
            QPushButton:hover    { background-color:#219653; }
            QPushButton:disabled { background-color:#95a5a6; }
        """)
        self.import_btn.clicked.connect(self.start_import)
        self.import_btn.setEnabled(False)
        layout.addWidget(self.import_btn)

        upd_btn = QtWidgets.QPushButton("⚖️  Update COG")
        upd_btn.setToolTip("Recalculate centres of gravity after repositioning")
        upd_btn.clicked.connect(self.update_cogs)
        layout.addWidget(upd_btn)

        clr_btn = QtWidgets.QPushButton("🗑️  Clear all")
        clr_btn.setStyleSheet("background-color:#e74c3c; color:white;")
        clr_btn.clicked.connect(self.clear_all)
        layout.addWidget(clr_btn)

        cls_btn = QtWidgets.QPushButton("✖️  Close")
        cls_btn.clicked.connect(self.close)
        layout.addWidget(cls_btn)

        widget.setLayout(layout)
        return widget

    # ======================================================================
    def load_defaults(self):
        ship = find_ship_object()
        if not ship:
            self.add_to_ship.setEnabled(False)
            self.status_bar.showMessage("⚠️  No Ship object found!", 5000)

    def browse_file(self):
        fp, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Select Excel file", "",
            "Excel Files (*.xlsx *.xls *.xlsm);;All Files (*.*)")
        if fp:
            self.file_path.setText(fp)
            self.load_sheets()
            if self.auto_preview.isChecked():
                self.preview_timer.start(1000)

    def load_sheets(self):
        fp = self.file_path.text()
        if not fp or not os.path.exists(fp):
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(fp, read_only=True)
            self.sheet_combo.clear()
            self.sheet_combo.addItem("(Automatic)", None)
            for n in wb.sheetnames:
                self.sheet_combo.addItem(n, n)
            wb.close()
            self.status_bar.showMessage(f"✓ {len(wb.sheetnames)} sheet(s) loaded")
        except Exception as e:
            self.status_bar.showMessage(f"❌ Error: {e}", 5000)

    def load_standard_preset(self):
        self.apply_config_to_ui(get_standard_config())
        self.status_bar.showMessage("✓ Standard layout loaded", 2000)

    def load_custom_preset(self):
        dlg = CustomConfigDialog(self)
        if dlg.exec_():
            self.apply_config_to_ui(dlg.get_config())
            self.status_bar.showMessage("✓ Custom layout loaded", 2000)

    def apply_config_to_ui(self, config):
        self.column_widgets['col_name'].setText(config.col_name)
        self.column_widgets['col_description'].setText(
            config.col_description or '')
        self.column_widgets['col_mass'].setText(config.col_mass)
        self.column_widgets['col_length'].setText(config.col_length)
        self.column_widgets['col_width'].setText(config.col_width)
        self.column_widgets['col_height'].setText(config.col_height)
        self.column_widgets['col_type'].setText(config.col_type or '')
        self.start_row.setValue(config.row_start)
        # 0 means "last filled row" (shows special text "Last filled row")
        self.end_row.setValue(config.row_end if config.row_end else 0)
        self.mass_unit.setCurrentIndex(
            {'kg':0,'t':1,'lbs':2}.get(config.mass_unit, 0))
        self.dim_unit.setCurrentIndex(
            {'m':0,'cm':1,'mm':2,'ft':3,'in':4}.get(config.dimension_unit, 0))
        self.default_desc.setText(config.default_description)
        self.default_type.setCurrentText(config.default_type)
        self.default_length.setValue(config.default_length)
        self.default_width.setValue(config.default_width)
        self.default_height.setValue(config.default_height)

    def get_config_from_ui(self):
        config = CargoImportConfig()
        config.col_name        = self.column_widgets['col_name'].text()   or 'A'
        desc                   = self.column_widgets['col_description'].text()
        config.col_description = desc if desc else None
        config.col_mass        = self.column_widgets['col_mass'].text()   or 'C'
        config.col_length      = self.column_widgets['col_length'].text() or 'D'
        config.col_width       = self.column_widgets['col_width'].text()  or 'E'
        config.col_height      = self.column_widgets['col_height'].text() or 'F'
        typ                    = self.column_widgets['col_type'].text()
        config.col_type        = typ if typ else None
        config.row_start       = self.start_row.value()
        end_val                = self.end_row.value()
        # 0 means "until last filled row"
        config.row_end         = end_val if end_val > 0 else None
        config.mass_unit       = self.mass_unit.currentText().split(' ')[0].lower()
        config.dimension_unit  = self.dim_unit.currentText().split(' ')[0].lower()
        config.default_description = self.default_desc.text()
        config.default_type        = self.default_type.currentText()
        config.default_length      = self.default_length.value()
        config.default_width       = self.default_width.value()
        config.default_height      = self.default_height.value()
        return config

    def on_column_changed(self):
        if self.auto_preview.isChecked():
            self.preview_timer.start(500)

    def update_preview(self):
        fp = self.file_path.text()
        if not fp or not os.path.exists(fp):
            self.preview_table.setRowCount(0)
            self.preview_count.setText("0 entries (file not found)")
            self.import_btn.setEnabled(False)
            return
        if not HAS_OPENPYXL:
            self.status_bar.showMessage("❌ openpyxl not installed!", 5000)
            return
        try:
            config     = self.get_config_from_ui()
            sheet_name = self.sheet_combo.currentData()
            if not sheet_name:
                sheet_name = self.sheet_name.text() or None

            import openpyxl
            wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
            ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active
            sheet_name = ws.title

            start_row    = config.row_start
            end_row_max  = config.row_end if config.row_end else min(ws.max_row, start_row + 49)
            col_name_idx = self._col_idx(config.col_name)
            col_mass_idx = self._col_idx(config.col_mass)

            preview_data = []
            for row in range(start_row, end_row_max + 1):
                nv = ws.cell(row=row, column=col_name_idx).value
                if not nv:
                    continue
                mv = ws.cell(row=row, column=col_mass_idx).value
                preview_data.append({
                    'row':   row,
                    'name':  str(nv),
                    'mass':  str(mv) if mv is not None else '—',
                    'valid': isinstance(mv, (int, float)) and mv > 0,
                })
            wb.close()

            self.preview_table.setRowCount(len(preview_data))
            for i, d in enumerate(preview_data):
                self.preview_table.setItem(
                    i, 0, QtWidgets.QTableWidgetItem(str(d['row'])))
                self.preview_table.setItem(
                    i, 1, QtWidgets.QTableWidgetItem(d['name']))
                mi = QtWidgets.QTableWidgetItem(d['mass'])
                if not d['valid']:
                    mi.setBackground(QtGui.QColor(255, 200, 200))
                    mi.setToolTip("Non-numeric value — this row will be skipped")
                self.preview_table.setItem(i, 2, mi)

            self.preview_count.setText(
                f"{len(preview_data)} entries found (sheet: {sheet_name})")
            self.preview_data = preview_data
            self.import_btn.setEnabled(
                any(d['valid'] for d in preview_data))
            self.status_bar.showMessage(
                f"✓ Preview updated — {len(preview_data)} entries", 2000)

        except Exception as e:
            self.status_bar.showMessage(f"❌ Error: {e}", 5000)
            import traceback; print(traceback.format_exc())

    def _col_idx(self, col):
        if isinstance(col, int): return col
        col = col.strip().upper(); r = 0
        for i, c in enumerate(reversed(col)):
            r += (ord(c) - ord('A') + 1) * (26 ** i)
        return r

    def start_import(self):
        fp = self.file_path.text()
        if not os.path.exists(fp):
            QtWidgets.QMessageBox.critical(self, "Error", "File not found!")
            return
        config     = self.get_config_from_ui()
        sheet_name = self.sheet_combo.currentData() or self.sheet_name.text() or None

        progress = QtWidgets.QProgressDialog(
            "Importing cargo…", "Cancel", 0, 100, self)
        progress.setWindowTitle("Import in progress…")
        progress.setWindowModality(QtCore.Qt.WindowModal)
        progress.show()
        try:
            progress.setValue(10)
            QtWidgets.QApplication.processEvents()
            self.created_weights = import_cargo_from_excel(
                filepath=fp, config=config, sheet_name=sheet_name,
                add_to_ship=self.add_to_ship.isChecked())
            progress.setValue(100)
            if self.created_weights:
                QtWidgets.QMessageBox.information(
                    self, "Success",
                    f"✓ {len(self.created_weights)} cargo object(s) imported!\n\n"
                    "• Objects can be repositioned in the 3D view\n"
                    "• Click 'Update COG' after repositioning")
                self.status_bar.showMessage(
                    f"✓ {len(self.created_weights)} object(s) imported", 5000)
            else:
                QtWidgets.QMessageBox.warning(
                    self, "Note",
                    "No cargo objects were imported.\n\n"
                    "Check the Python console for details — common causes:\n"
                    "• 'From row' includes the header row (try row 2)\n"
                    "• Column letters don't match your spreadsheet\n"
                    "• Mass values are 0 or contain text")
        except Exception as e:
            QtWidgets.QMessageBox.critical(
                self, "Error", f"Import failed:\n{e}")
            import traceback; print(traceback.format_exc())
        finally:
            progress.close()

    def update_cogs(self):
        try:
            update_all_cargo_cogs()
            QtWidgets.QMessageBox.information(
                self, "Success", "✓ COG updated for all cargo objects")
            self.status_bar.showMessage("✓ COGs updated", 3000)
        except Exception as e:
            QtWidgets.QMessageBox.critical(
                self, "Error", f"COG update failed:\n{e}")

    def clear_all(self):
        if not self.created_weights:
            QtWidgets.QMessageBox.information(
                self, "Info", "No objects to delete.")
            return
        reply = QtWidgets.QMessageBox.question(
            self, "Confirm",
            f"Delete all {len(self.created_weights)} imported cargo object(s)?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        if reply == QtWidgets.QMessageBox.Yes:
            doc = App.ActiveDocument
            if doc:
                ship = find_ship_object()
                if ship and hasattr(ship, 'Weights'):
                    wl = list(ship.Weights)
                    for w in self.created_weights:
                        if w.Name in wl: wl.remove(w.Name)
                    ship.Weights = wl
                for w in self.created_weights:
                    doc.removeObject(w.Name)
                doc.recompute()
                count = len(self.created_weights)
                self.created_weights = []
                self.status_bar.showMessage(f"✓ {count} object(s) deleted", 3000)
                QtWidgets.QMessageBox.information(
                    self, "Done", "All objects deleted.")


# ============================================================================
class CustomConfigDialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Custom Layout")
        layout = QtWidgets.QVBoxLayout()
        layout.addWidget(QtWidgets.QLabel(
            "Enter the column letters for your layout:"))
        form = QtWidgets.QGridLayout()
        fields = [('Name','A'),('Mass','C'),('Length','D'),
                  ('Width','E'),('Height','F')]
        self._edits = {}
        for i, (lbl, default) in enumerate(fields):
            form.addWidget(QtWidgets.QLabel(f"{lbl}:"), i, 0)
            le = QtWidgets.QLineEdit(default)
            form.addWidget(le, i, 1)
            self._edits[lbl.lower()] = le
        layout.addLayout(form)
        btns = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)
        self.setLayout(layout)

    def get_config(self):
        return get_custom_config(
            name_col  = self._edits['name'].text(),
            mass_col  = self._edits['mass'].text(),
            length_col= self._edits['length'].text(),
            width_col = self._edits['width'].text(),
            height_col= self._edits['height'].text(),
        )


# ============================================================================
class CargoImportCommand:
    def GetResources(self):
        return {'Pixmap':'freecad',
                'MenuText':'Cargo Stowage Plan Import',
                'ToolTip':'Import cargo from an Excel spreadsheet',
                'Accel':'Ctrl+Shift+C'}

    def Activated(self):
        if not HAS_OPENPYXL:
            QtWidgets.QMessageBox.critical(
                None, "Missing dependency",
                "openpyxl is not installed!\n\n"
                "Install it in the FreeCAD Python console:\n"
                "  import subprocess, sys\n"
                "  subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl'])")
            return
        dialog = CargoImportDialog(Gui.getMainWindow())
        dialog.exec_()

    def IsActive(self):
        return App.ActiveDocument is not None


if __name__ == "__main__":
    dialog = CargoImportDialog()
    dialog.exec_()
else:
    Gui.addCommand('CargoImportCommand', CargoImportCommand())
